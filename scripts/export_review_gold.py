from __future__ import annotations

import argparse
import atexit
import csv
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REVIEWED_WORKBOOK = REPO_ROOT / "activity_asset_review_GP First Pass.xlsx"
DEFAULT_ORIGINAL_WORKBOOK = REPO_ROOT / "activity_asset_review.xlsx"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "data" / "gold"
SHEET_NAME = "activity_asset_review"
REVIEWED_DATA_ROWS = 165

CANDIDATE_REVIEW_FILENAME = "activity_asset_gold_candidate_review.csv"
REQUIREMENTS_FILENAME = "activity_asset_gold_requirements_multiasset.csv"
ACTIVITY_SUMMARY_FILENAME = "activity_asset_gold_activity_multiasset_summary.csv"
MULTILABEL_WIDE_FILENAME = "activity_asset_gold_multilabel_wide.csv"
SUMMARY_FILENAME = "activity_asset_gold_export_summary.json"

ASSET_TYPES = [
    "crane",
    "hoist",
    "loading_bay",
    "ewp",
    "concrete_pump",
    "excavator",
    "forklift",
    "telehandler",
    "compactor",
]

SLOTS = {
    1: {"type": 10, "confidence": 11, "role": 12, "estimated_hours": 13, "profile_shape": 14},
    2: {"type": 15, "confidence": 16, "role": 17, "estimated_hours": 18, "profile_shape": 19},
    3: {"type": 20, "confidence": 21, "role": 22, "estimated_hours": 23, "profile_shape": 24},
    4: {"type": 25, "confidence": 26, "role": 27, "estimated_hours": 28, "profile_shape": 29},
}

PARENS_RE = re.compile(r"^(.*?)\s*\(([^()]*)\)\s*$")


def _cell_value(ws, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def _changed(original: Any, reviewed: Any) -> bool:
    return original != reviewed


def _reviewed_token(original: Any, reviewed: Any) -> Any:
    if _changed(original, reviewed) and isinstance(reviewed, str):
        match = PARENS_RE.match(reviewed.strip())
        if match:
            return match.group(2).strip()
    return reviewed


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _canonical_asset_type(value: Any) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    key = re.sub(r"[\s\-]+", "_", text.lower())
    aliases = {
        "loading_bay": "loading_bay",
        "loadingbay": "loading_bay",
        "loading_bays": "loading_bay",
        "concrete_pump": "concrete_pump",
        "concretepump": "concrete_pump",
        "ewp": "ewp",
        "e_w_p": "ewp",
    }
    return aliases.get(key, key)


def _canonical_role(value: Any) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    return re.sub(r"[\s\-]+", "_", text.lower())


def _canonical_shape(value: Any) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    return re.sub(r"[\s\-]+", "_", text.lower())


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _confidence_tier(value: float | None) -> str | None:
    if value is None or value <= 0:
        return None
    if value >= 0.75:
        return "high"
    if value >= 0.40:
        return "medium"
    return "low"


def _float_changed(a: float | None, b: float | None) -> bool:
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    return not math.isclose(a, b, rel_tol=1e-9, abs_tol=1e-9)


def _slot_record(original_ws, reviewed_ws, row: int, slot: int) -> dict[str, Any] | None:
    cols = SLOTS[slot]
    generated = {
        field: _cell_value(original_ws, row, col)
        for field, col in cols.items()
    }
    reviewed_raw = {
        field: _cell_value(reviewed_ws, row, col)
        for field, col in cols.items()
    }
    if not any(generated.values()) and not any(reviewed_raw.values()):
        return None

    generated_clean = {
        "type": _canonical_asset_type(generated["type"]),
        "confidence": _to_float(generated["confidence"]),
        "role": _canonical_role(generated["role"]),
        "estimated_hours": _to_float(generated["estimated_hours"]),
        "profile_shape": _canonical_shape(generated["profile_shape"]),
    }
    reviewed_clean = {
        "type": _canonical_asset_type(_reviewed_token(generated["type"], reviewed_raw["type"])),
        "confidence": _to_float(_reviewed_token(generated["confidence"], reviewed_raw["confidence"])),
        "role": _canonical_role(_reviewed_token(generated["role"], reviewed_raw["role"])),
        "estimated_hours": _to_float(_reviewed_token(generated["estimated_hours"], reviewed_raw["estimated_hours"])),
        "profile_shape": _canonical_shape(_reviewed_token(generated["profile_shape"], reviewed_raw["profile_shape"])),
    }

    changed_fields: list[str] = []
    for field in ("type", "role", "profile_shape"):
        if generated_clean[field] != reviewed_clean[field]:
            changed_fields.append(field)
    for field in ("confidence", "estimated_hours"):
        if _float_changed(generated_clean[field], reviewed_clean[field]):
            changed_fields.append(field)

    reviewed_required = bool(
        reviewed_clean["type"]
        and reviewed_clean["type"] != "none"
        and reviewed_clean["confidence"] is not None
        and reviewed_clean["confidence"] > 0
    )
    if not reviewed_required:
        review_status = "dropped"
    elif changed_fields:
        review_status = "corrected"
    else:
        review_status = "accepted"

    label_hours = reviewed_clean["estimated_hours"] if reviewed_required else 0.0

    return {
        "candidate_slot": slot,
        "generated_asset_type": generated_clean["type"],
        "generated_confidence": generated_clean["confidence"],
        "generated_role": generated_clean["role"],
        "generated_estimated_hours": generated_clean["estimated_hours"],
        "generated_profile_shape": generated_clean["profile_shape"],
        "reviewed_asset_type": reviewed_clean["type"],
        "reviewed_confidence": reviewed_clean["confidence"],
        "reviewed_confidence_tier": _confidence_tier(reviewed_clean["confidence"]),
        "reviewed_role": reviewed_clean["role"],
        "reviewed_estimated_hours": reviewed_clean["estimated_hours"],
        "reviewed_profile_shape": reviewed_clean["profile_shape"],
        "reviewed_is_required": reviewed_required,
        "label_estimated_hours": label_hours,
        "review_status": review_status,
        "changed_fields": "|".join(changed_fields),
    }


def _write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _display_path(path: Path) -> str:
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export reviewed multi-asset activity gold labels from the review workbook.",
    )
    parser.add_argument(
        "--original-workbook",
        type=Path,
        default=DEFAULT_ORIGINAL_WORKBOOK,
        help="Path to the generated/source review workbook.",
    )
    parser.add_argument(
        "--reviewed-workbook",
        type=Path,
        default=DEFAULT_REVIEWED_WORKBOOK,
        help="Path to the manually reviewed workbook.",
    )
    parser.add_argument(
        "--sheet-name",
        default=SHEET_NAME,
        help="Worksheet name containing the review table.",
    )
    parser.add_argument(
        "--reviewed-rows",
        type=int,
        default=REVIEWED_DATA_ROWS,
        help="Number of reviewed activity rows to export.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory where multi-asset gold CSV/JSON files will be written.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    candidate_review_csv = output_dir / CANDIDATE_REVIEW_FILENAME
    requirements_csv = output_dir / REQUIREMENTS_FILENAME
    activity_summary_csv = output_dir / ACTIVITY_SUMMARY_FILENAME
    multilabel_wide_csv = output_dir / MULTILABEL_WIDE_FILENAME
    summary_json = output_dir / SUMMARY_FILENAME

    original_wb = load_workbook(args.original_workbook, data_only=False)
    reviewed_wb = load_workbook(args.reviewed_workbook, data_only=False)
    atexit.register(original_wb.close)
    atexit.register(reviewed_wb.close)
    original_ws = original_wb[args.sheet_name]
    reviewed_ws = reviewed_wb[args.sheet_name]

    candidate_rows: list[dict[str, Any]] = []
    requirement_rows: list[dict[str, Any]] = []
    positive_by_activity: dict[int, list[dict[str, Any]]] = defaultdict(list)
    dropped_by_activity: dict[int, list[dict[str, Any]]] = defaultdict(list)
    activity_rows: dict[int, dict[str, Any]] = {}

    for row in range(2, args.reviewed_rows + 2):
        activity = {
            "activity_number": _cell_value(reviewed_ws, row, 1),
            "activity_name": _cell_value(reviewed_ws, row, 2),
            "activity_hierarchy": _cell_value(reviewed_ws, row, 3),
            "activity_name_normalized": _cell_value(reviewed_ws, row, 4),
            "duration_days": _cell_value(reviewed_ws, row, 5),
            "profile_days_used": _cell_value(reviewed_ws, row, 6),
            "asset_candidate_count": _cell_value(reviewed_ws, row, 7),
            "review_priority": _cell_value(reviewed_ws, row, 8),
            "outlier_flags": _cell_value(reviewed_ws, row, 9),
            "review_source": "manual_first_pass",
        }
        try:
            activity_number = int(activity["activity_number"])
        except (TypeError, ValueError):
            print(f"Skipping review row {row}: invalid activity_number={activity['activity_number']!r}")
            continue
        activity_rows[activity_number] = activity

        for slot in SLOTS:
            slot_row = _slot_record(original_ws, reviewed_ws, row, slot)
            if slot_row is None:
                continue

            candidate_row = {**activity, **slot_row}
            candidate_rows.append(candidate_row)

            if slot_row["reviewed_is_required"]:
                requirement_row = {
                    "activity_number": activity["activity_number"],
                    "activity_name": activity["activity_name"],
                    "activity_hierarchy": activity["activity_hierarchy"],
                    "activity_name_normalized": activity["activity_name_normalized"],
                    "duration_days": activity["duration_days"],
                    "profile_days_used": activity["profile_days_used"],
                    "asset_type": slot_row["reviewed_asset_type"],
                    "label_confidence": slot_row["reviewed_confidence"],
                    "confidence_tier": slot_row["reviewed_confidence_tier"],
                    "asset_role": slot_row["reviewed_role"],
                    "estimated_total_hours": slot_row["label_estimated_hours"],
                    "profile_shape": slot_row["reviewed_profile_shape"],
                    "candidate_slot": slot,
                    "is_lead_asset": slot_row["reviewed_role"] == "lead",
                    "is_support_asset": slot_row["reviewed_role"] == "support",
                    "is_incidental_asset": slot_row["reviewed_role"] == "incidental",
                    "review_status": slot_row["review_status"],
                    "changed_fields": slot_row["changed_fields"],
                    "review_source": activity["review_source"],
                }
                requirement_rows.append(requirement_row)
                positive_by_activity[activity_number].append(requirement_row)
            else:
                dropped_by_activity[activity_number].append(candidate_row)

    activity_summary_rows: list[dict[str, Any]] = []
    multilabel_rows: list[dict[str, Any]] = []
    role_rank = {"lead": 0, "support": 1, "incidental": 2, None: 9}
    for activity_number in sorted(activity_rows):
        activity = activity_rows[activity_number]
        positives = positive_by_activity.get(activity_number, [])
        dropped = dropped_by_activity.get(activity_number, [])
        sorted_positives = sorted(
            positives,
            key=lambda r: (
                role_rank.get(r["asset_role"], 9),
                int(r["candidate_slot"]),
                str(r["asset_type"] or ""),
            ),
        )
        if positives:
            lead_assets = [r["asset_type"] for r in sorted_positives if r["asset_role"] == "lead"]
            support_assets = [r["asset_type"] for r in sorted_positives if r["asset_role"] == "support"]
            incidental_assets = [r["asset_type"] for r in sorted_positives if r["asset_role"] == "incidental"]
        else:
            lead_assets = []
            support_assets = []
            incidental_assets = []

        activity_summary_rows.append({
            "activity_number": activity["activity_number"],
            "activity_name": activity["activity_name"],
            "activity_hierarchy": activity["activity_hierarchy"],
            "activity_name_normalized": activity["activity_name_normalized"],
            "duration_days": activity["duration_days"],
            "profile_days_used": activity["profile_days_used"],
            "asset_requirement_count": len(positives),
            "lead_asset_count": len(lead_assets),
            "support_asset_count": len(support_assets),
            "incidental_asset_count": len(incidental_assets),
            "reviewed_asset_types": "|".join(str(r["asset_type"]) for r in sorted_positives),
            "lead_asset_types": "|".join(str(v) for v in lead_assets),
            "support_asset_types": "|".join(str(v) for v in support_assets),
            "incidental_asset_types": "|".join(str(v) for v in incidental_assets),
            "total_estimated_hours_all_assets": round(
                sum(float(r["estimated_total_hours"] or 0) for r in positives),
                4,
            ),
            "dropped_candidate_count": len(dropped),
            "dropped_generated_asset_types": "|".join(
                str(r["generated_asset_type"]) for r in dropped if r["generated_asset_type"]
            ),
            "has_multiple_assets": len(positives) > 1,
            "has_no_assets": len(positives) == 0,
            "review_source": activity["review_source"],
        })

        wide = {
            "activity_number": activity["activity_number"],
            "activity_name": activity["activity_name"],
            "activity_hierarchy": activity["activity_hierarchy"],
            "activity_name_normalized": activity["activity_name_normalized"],
            "duration_days": activity["duration_days"],
            "profile_days_used": activity["profile_days_used"],
            "asset_requirement_count": len(positives),
            "review_source": activity["review_source"],
        }
        by_asset = {r["asset_type"]: r for r in positives}
        for asset_type in ASSET_TYPES:
            row = by_asset.get(asset_type)
            prefix = asset_type
            wide[f"{prefix}_required"] = bool(row)
            wide[f"{prefix}_confidence"] = row["label_confidence"] if row else ""
            wide[f"{prefix}_role"] = row["asset_role"] if row else ""
            wide[f"{prefix}_estimated_hours"] = row["estimated_total_hours"] if row else 0.0
            wide[f"{prefix}_profile_shape"] = row["profile_shape"] if row else ""
        multilabel_rows.append(wide)

    candidate_fields = [
        "activity_number",
        "activity_name",
        "activity_hierarchy",
        "activity_name_normalized",
        "duration_days",
        "profile_days_used",
        "asset_candidate_count",
        "review_priority",
        "outlier_flags",
        "candidate_slot",
        "generated_asset_type",
        "generated_confidence",
        "generated_role",
        "generated_estimated_hours",
        "generated_profile_shape",
        "reviewed_asset_type",
        "reviewed_confidence",
        "reviewed_confidence_tier",
        "reviewed_role",
        "reviewed_estimated_hours",
        "reviewed_profile_shape",
        "reviewed_is_required",
        "label_estimated_hours",
        "review_status",
        "changed_fields",
        "review_source",
    ]
    requirement_fields = [
        "activity_number",
        "activity_name",
        "activity_hierarchy",
        "activity_name_normalized",
        "duration_days",
        "profile_days_used",
        "asset_type",
        "label_confidence",
        "confidence_tier",
        "asset_role",
        "estimated_total_hours",
        "profile_shape",
        "candidate_slot",
        "is_lead_asset",
        "is_support_asset",
        "is_incidental_asset",
        "review_status",
        "changed_fields",
        "review_source",
    ]
    summary_fields = [
        "activity_number",
        "activity_name",
        "activity_hierarchy",
        "activity_name_normalized",
        "duration_days",
        "profile_days_used",
        "asset_requirement_count",
        "lead_asset_count",
        "support_asset_count",
        "incidental_asset_count",
        "reviewed_asset_types",
        "lead_asset_types",
        "support_asset_types",
        "incidental_asset_types",
        "total_estimated_hours_all_assets",
        "dropped_candidate_count",
        "dropped_generated_asset_types",
        "has_multiple_assets",
        "has_no_assets",
        "review_source",
    ]
    multilabel_fields = [
        "activity_number",
        "activity_name",
        "activity_hierarchy",
        "activity_name_normalized",
        "duration_days",
        "profile_days_used",
        "asset_requirement_count",
        "review_source",
    ]
    for asset_type in ASSET_TYPES:
        multilabel_fields.extend([
            f"{asset_type}_required",
            f"{asset_type}_confidence",
            f"{asset_type}_role",
            f"{asset_type}_estimated_hours",
            f"{asset_type}_profile_shape",
        ])

    _write_csv(candidate_review_csv, candidate_fields, candidate_rows)
    _write_csv(requirements_csv, requirement_fields, requirement_rows)
    _write_csv(activity_summary_csv, summary_fields, activity_summary_rows)
    _write_csv(multilabel_wide_csv, multilabel_fields, multilabel_rows)

    summary = {
        "reviewed_activity_rows": args.reviewed_rows,
        "candidate_rows": len(candidate_rows),
        "positive_requirement_rows": len(requirement_rows),
        "activity_summary_rows": len(activity_summary_rows),
        "multilabel_wide_rows": len(multilabel_rows),
        "multi_asset_activity_rows": sum(1 for row in activity_summary_rows if row["has_multiple_assets"]),
        "single_asset_activity_rows": sum(
            1 for row in activity_summary_rows if row["asset_requirement_count"] == 1
        ),
        "no_asset_activity_rows": sum(1 for row in activity_summary_rows if row["has_no_assets"]),
        "candidate_status_counts": dict(Counter(row["review_status"] for row in candidate_rows)),
        "positive_asset_type_counts": dict(Counter(row["asset_type"] for row in requirement_rows)),
        "asset_role_counts": dict(Counter(row["asset_role"] for row in requirement_rows)),
        "dropped_generated_asset_type_counts": dict(
            Counter(row["generated_asset_type"] for row in candidate_rows if row["review_status"] == "dropped")
        ),
        "changed_field_counts": dict(
            Counter(
                field
                for row in candidate_rows
                for field in str(row["changed_fields"] or "").split("|")
                if field
            )
        ),
        "outputs": [
            _display_path(requirements_csv),
            _display_path(candidate_review_csv),
            _display_path(activity_summary_csv),
            _display_path(multilabel_wide_csv),
            _display_path(summary_json),
        ],
    }
    summary_json.write_text(json.dumps(summary, indent=2, sort_keys=True), encoding="utf-8")
    print(json.dumps(summary, indent=2, sort_keys=True))
    original_wb.close()
    reviewed_wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
