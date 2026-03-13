"""
Standalone test script for the AI service.
Run from the repo root:

    python -m scripts.test_ai_service --file path/to/your_programme.csv
    python -m scripts.test_ai_service --file path/to/your_programme.xlsx
    python -m scripts.test_ai_service --demo   (runs with built-in demo data)

What it checks:
  1. API key is working
  2. Structure detection returns valid column_mapping + ActivityItem list
  3. Classification returns valid asset_types for sample activities
  4. Keyword pre-screening and confidence tiers work correctly
  5. Fallback chain works when AI is disabled (AI_ENABLED=false)
  6. Subcontractor asset suggestion heuristic
"""

import asyncio
import csv
import os
import sys
from pathlib import Path

# Make sure we can import from app/
sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv
load_dotenv()


def load_csv(filepath: str) -> list[dict]:
    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [row for row in reader]


def load_xlsx(filepath: str) -> list[dict]:
    if filepath.lower().endswith(".xls"):
        raise ValueError(f"Legacy .xls format is not supported. Convert '{filepath}' to .xlsx first.")
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [
        str(cell.value).strip() if cell.value else f"col_{i}"
        for i, cell in enumerate(ws[1])
    ]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: val for i, val in enumerate(row)})
    wb.close()
    return rows


DEMO_ROWS = [
    {"Activity ID": "A1000", "Activity Name": "Bulk Excavation Zone A", "Original Duration": "15d", "Start": "03/03/2026", "Finish": "21/03/2026", "Outline Level": "2", "Resource Names": "Earthworks Sub"},
    {"Activity ID": "A1010", "Activity Name": "Install Piling", "Original Duration": "20d", "Start": "23/03/2026", "Finish": "17/04/2026", "Outline Level": "2", "Resource Names": "Piling Sub"},
    {"Activity ID": "A1020", "Activity Name": "Pour Pad Footings", "Original Duration": "5d", "Start": "20/04/2026", "Finish": "24/04/2026", "Outline Level": "2", "Resource Names": "Concreter"},
    {"Activity ID": "A2000", "Activity Name": "Level 1-4 Structural Steel Erection", "Original Duration": "30d", "Start": "27/04/2026", "Finish": "05/06/2026", "Outline Level": "2", "Resource Names": "Steel Fixer Sub"},
    {"Activity ID": "A2010", "Activity Name": "Level 1-4 Precast Panels Install", "Original Duration": "25d", "Start": "08/06/2026", "Finish": "10/07/2026", "Outline Level": "2", "Resource Names": "Precast Sub"},
    {"Activity ID": "A2020", "Activity Name": "Level 5-8 Slab Formwork + Pour", "Original Duration": "20d", "Start": "13/07/2026", "Finish": "07/08/2026", "Outline Level": "2", "Resource Names": "Concreter"},
    {"Activity ID": "A3000", "Activity Name": "Electrical Conduit Rough-In Levels 1-4", "Original Duration": "40d", "Start": "10/08/2026", "Finish": "04/09/2026", "Outline Level": "3", "Resource Names": "Electrical Sub"},
    {"Activity ID": "A3010", "Activity Name": "HVAC Ductwork Install Levels 1-4", "Original Duration": "35d", "Start": "07/09/2026", "Finish": "23/10/2026", "Outline Level": "3", "Resource Names": "Mechanical Sub"},
    {"Activity ID": "A4000", "Activity Name": "External Cladding + Facade", "Original Duration": "60d", "Start": "26/10/2026", "Finish": "15/01/2027", "Outline Level": "2", "Resource Names": "Cladding Sub"},
    {"Activity ID": "A9000", "Activity Name": "Practical Completion", "Original Duration": "0d", "Start": "15/03/2027", "Finish": "15/03/2027", "Outline Level": "1", "Resource Names": ""},
]

DEMO_ACTIVITIES = [
    {"id": "A1000", "name": "Bulk Excavation Zone A"},
    {"id": "A1010", "name": "Install Piling"},
    {"id": "A1020", "name": "Pour Pad Footings"},
    {"id": "A2000", "name": "Level 1-4 Structural Steel Erection"},
    {"id": "A2010", "name": "Level 1-4 Precast Panels Install"},
    {"id": "A2020", "name": "Level 5-8 Slab Formwork + Pour"},
    {"id": "A3000", "name": "Electrical Conduit Rough-In Levels 1-4"},
    {"id": "A3010", "name": "HVAC Ductwork Install Levels 1-4"},
    {"id": "A4000", "name": "External Cladding + Facade"},
    {"id": "A9000", "name": "Practical Completion"},
]

DEMO_SUBCONTRACTORS = [
    {"id": "sub-001", "trade_specialty": "earthworks"},
    {"id": "sub-002", "trade_specialty": "structural"},
    {"id": "sub-003", "trade_specialty": "concreter"},
    {"id": "sub-004", "trade_specialty": "electrician"},
    {"id": "sub-005", "trade_specialty": "mechanical"},
    {"id": "sub-006", "trade_specialty": "facade"},
]


def print_section(title: str):
    print(f"\n{'─'*60}")
    print(f"  {title}")
    print(f"{'─'*60}")


async def run_structure_test(rows: list[dict]):
    from app.services.ai_service import detect_structure, ALLOWED_ASSET_TYPES

    print_section("TEST 1: Structure Detection")
    print(f"  Input: {len(rows)} rows, {len(rows[0]) if rows else 0} columns")
    print(f"  Headers: {list(rows[0].keys()) if rows else []}")

    try:
        result = await detect_structure(rows[:100])

        print(f"\n  completeness_score : {result.completeness_score}")
        print(f"  activities parsed  : {len(result.activities)}")
        print(f"  missing_fields     : {result.missing_fields}")
        print(f"  notes              : {result.notes}")

        print("\n  Column Mapping:")
        for field_name, col_header in result.column_mapping.items():
            print(f"    ✅  {field_name:20s} → {col_header}")

        if result.activities:
            print("\n  First 3 activities:")
            for act in result.activities[:3]:
                print(f"    id={act.id}  name={act.name!r}  start={act.start}  finish={act.finish}  summary={act.is_summary}")

        return result

    except Exception as e:
        print(f"\n  ❌ Error: {e}")
        raise


async def run_classification_test(activities: list[dict]):
    from app.services.ai_service import classify_assets, ALLOWED_ASSET_TYPES

    print_section("TEST 2: Asset Classification")
    print(f"  Input: {len(activities)} activities")

    try:
        result = await classify_assets(activities)

        print(f"\n  Total classified   : {len(result.classifications)}")
        print(f"  Skipped (low conf) : {len(result.skipped)}")
        print(f"  Tokens used        : {result.batch_tokens_used}")

        high = [c for c in result.classifications if c.confidence == "high"]
        med = [c for c in result.classifications if c.confidence == "medium"]
        low_kept = [c for c in result.classifications if c.confidence == "low"]
        print(f"\n  Confidence: high={len(high)}  medium={len(med)}  low(kept)={len(low_kept)}  skipped={len(result.skipped)}")

        print("\n  Classifications:")
        for c in result.classifications:
            icon = {"high": "🟢", "medium": "🟡", "low": "🔴"}.get(c.confidence, "⚪")
            valid = "✅" if c.asset_type in ALLOWED_ASSET_TYPES else "❌ INVALID"
            print(f"    {icon} [{c.confidence:6s}] {c.asset_type:15s} {valid}  src={c.source:12s}  id={c.activity_id}")
            if c.reasoning:
                print(f"           ↳ {c.reasoning}")

        if result.skipped:
            print(f"\n  Skipped IDs: {result.skipped}")

        invalid = [c for c in result.classifications if c.asset_type not in ALLOWED_ASSET_TYPES]
        if invalid:
            print(f"\n  ❌ VALIDATION FAILED — {len(invalid)} invalid asset_type values!")
        else:
            print("\n  ✅ All asset_type values are valid")

    except Exception as e:
        print(f"\n  ❌ Error: {e}")
        raise


def run_sub_suggestion_test():
    from app.services.ai_service import suggest_subcontractor_asset_types, ALLOWED_ASSET_TYPES

    print_section("TEST 3: Subcontractor Asset Suggestions")
    print(f"  Input: {len(DEMO_SUBCONTRACTORS)} subcontractors")

    results = suggest_subcontractor_asset_types(DEMO_SUBCONTRACTORS)
    for suggestion in results:
        valid = all(t in ALLOWED_ASSET_TYPES or t == "other" for t in suggestion.suggested_asset_types)
        icon = "✅" if valid else "❌"
        print(f"  {icon}  [{suggestion.trade_specialty:20s}]  →  {suggestion.suggested_asset_types}")


async def run_fallback_test():
    print_section("TEST 4: AI_ENABLED=false fallback")
    original = os.environ.get("AI_ENABLED")
    os.environ["AI_ENABLED"] = "false"

    import importlib
    import app.core.config as cfg_module
    importlib.reload(cfg_module)
    import app.services.ai_service as ai_module
    importlib.reload(ai_module)

    result_struct = await ai_module.detect_structure(DEMO_ROWS[:10])
    print(f"  detect_structure fallback: completeness={result_struct.completeness_score}, activities={len(result_struct.activities)}")
    print("  ✅ Returned StructureResult (no exception)")

    result_class = await ai_module.classify_assets(DEMO_ACTIVITIES[:5])
    print(f"  classify_assets fallback: classified={len(result_class.classifications)}, skipped={len(result_class.skipped)}, tokens={result_class.batch_tokens_used}")
    print("  ✅ Returned ClassificationResult (no exception)")

    if original is None:
        del os.environ["AI_ENABLED"]
    else:
        os.environ["AI_ENABLED"] = original
    importlib.reload(cfg_module)
    importlib.reload(ai_module)


async def run_pdf_test(filepath: Path):
    """Test the full PDF extraction pipeline on a real PDF file."""
    from app.services.ai_service import (
        extract_pdf_activities,
        parse_ms_project_pdf,
    )

    print_section("PDF EXTRACTION TEST")
    print(f"  File: {filepath.name} ({filepath.stat().st_size // 1024} KB)")

    pdf_bytes = filepath.read_bytes()
    rows: list = []

    # Stage 1: MS Project / P6 text-regex parser (fastest, no AI, handles Gantt PDFs)
    print("\n  Stage 1: MS Project PDF text-regex parser...")
    rows, error = parse_ms_project_pdf(pdf_bytes)
    if error:
        print(f"  ❌ MS Project parser error: {error}")
    elif rows:
        summary_count = sum(1 for r in rows if r.get("Is Summary") == "Yes")
        milestone_count = sum(1 for r in rows if r.get("Is Milestone") == "Yes")
        detail_count = len(rows) - summary_count - milestone_count
        print(f"  ✅ MS Project parser extracted {len(rows)} rows "
              f"({detail_count} detail, {summary_count} summary, {milestone_count} milestones)")
        print(f"  Headers: {list(rows[0].keys())}")
        print(f"  Sample rows:")
        for r in rows[:4]:
            flag = " [SUMMARY]" if r.get("Is Summary") == "Yes" else ""
            print(f"    [{r.get('ID'):>5}] {r.get('Name', '')[:55]}{flag}")
    else:
        print("  ⚠️  MS Project parser found no rows — trying pdfplumber...")

    # Stage 2: pdfplumber (for other structured PDF table formats)
    if not rows:
        print("\n  Stage 2: pdfplumber table extraction...")
        try:
            import pdfplumber as _pdfplumber
            import io as _io
            with _pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
                tables = []
                for page in pdf.pages:
                    t = page.extract_table()
                    if t:
                        tables.extend(t)
                rows = tables  # type: ignore[assignment]
                error = None
        except Exception as exc:
            rows = []
            error = str(exc)
        if error:
            print(f"  ❌ pdfplumber error: {error}")
        elif rows:
            print(f"  ✅ pdfplumber found {len(rows)} rows")
        else:
            print("  ⚠️  pdfplumber found no tables — will try Claude Vision")

    # Stage 3: Claude Vision (last resort for image/scanned PDFs)
    if not rows:
        print("\n  Stage 3: Claude Vision extraction...")
        try:
            rows = await extract_pdf_activities(pdf_bytes)
            print(f"  ✅ Claude Vision extracted {len(rows)} rows")
        except Exception as e:
            print(f"  ❌ Claude Vision failed: {e}")
            return

    if not rows:
        print("  ❌ No rows extracted from PDF — cannot continue")
        return

    # Stage 4: Structure detection
    print(f"\n  Stage 4: Structure detection on {len(rows)} rows...")
    from app.services.ai_service import detect_structure
    result = await detect_structure(rows)
    print(f"  completeness_score : {result.completeness_score}")
    print(f"  activities parsed  : {len(result.activities)}")
    print(f"  Column Mapping     :")
    for field, col in result.column_mapping.items():
        print(f"    {field:20s} → {col}")
    if result.notes:
        print(f"  notes              : {result.notes}")

    # Stage 5: Classify first 30 non-summary activities
    if result.activities:
        non_summary = [a for a in result.activities if not a.is_summary][:30]
        print(f"\n  Stage 5: Classifying {len(non_summary)} detail activities (first 30)...")
        from app.services.ai_service import classify_assets
        act_dicts = [{"id": a.id, "name": a.name} for a in non_summary]
        classification = await classify_assets(act_dicts)
        print(f"  Classified: {len(classification.classifications)}  Skipped (low-conf): {len(classification.skipped)}")
        print(f"  Asset type breakdown:")
        asset_counts: dict = {}
        for c in classification.classifications:
            asset_counts[c.asset_type] = asset_counts.get(c.asset_type, 0) + 1
        for k, v in sorted(asset_counts.items(), key=lambda x: -x[1]):
            print(f"    {k:20s}: {v}")
        print(f"\n  Sample classifications:")
        for c in classification.classifications[:15]:
            icon = {"high": "✅", "medium": "⚡", "low": "⚠️ "}.get(c.confidence, "  ")
            act_name = next((a.name for a in non_summary if str(a.id) == str(c.activity_id)), c.activity_id)
            print(f"    {icon} [{c.confidence:6s}] {c.asset_type:15s}  {act_name[:55]}")


async def main():
    import argparse
    parser = argparse.ArgumentParser(description="Test the SiteSpace AI service")
    parser.add_argument("--file", help="Path to CSV, XLSX, or PDF programme file")
    parser.add_argument("--demo", action="store_true", help="Use built-in demo data")
    parser.add_argument("--classify-only", action="store_true", help="Skip structure, run classification only")
    parser.add_argument("--sub-only", action="store_true", help="Run subcontractor suggestion test only")
    args = parser.parse_args()

    print("\n🔧 SiteSpace AI Service — Test Runner")

    from app.core.config import settings
    ai_status = "ENABLED" if settings.AI_ENABLED else "DISABLED (fallback mode)"
    has_key = "✅ configured" if settings.AI_API_KEY else "❌ NOT SET"
    print(f"   AI status  : {ai_status}")
    print(f"   AI_API_KEY : {has_key}")
    print(f"   AI_MODEL   : {settings.AI_MODEL}")

    if args.sub_only:
        run_sub_suggestion_test()
        return

    if args.file:
        filepath = Path(args.file)
        if not filepath.exists():
            print(f"❌ File not found: {filepath}")
            sys.exit(1)
        ext = filepath.suffix.lower()

        # PDF gets its own dedicated test path
        if ext == ".pdf":
            await run_pdf_test(filepath)
            run_sub_suggestion_test()
            return

        if ext == ".csv":
            rows = load_csv(str(filepath))
        elif ext in (".xlsx", ".xlsm", ".xls"):
            rows = load_xlsx(str(filepath))
        else:
            print(f"❌ Unsupported file type: {ext}. Use .csv, .xlsx, or .pdf")
            sys.exit(1)
        print(f"   File: {filepath.name} ({len(rows)} rows)")
        activities = [
            {"id": str(i), "name": str(r.get(list(r.keys())[0], ""))}
            for i, r in enumerate(rows[:50])
        ]
    else:
        print("   Mode: demo data (use --file path/to/programme.csv for real data)")
        rows = DEMO_ROWS
        activities = DEMO_ACTIVITIES

    if not args.classify_only:
        await run_structure_test(rows)

    await run_classification_test(activities)
    run_sub_suggestion_test()
    await run_fallback_test()

    print(f"\n{'─'*60}")
    print("  Done. Review output above for any ❌ failures.")
    print(f"{'─'*60}\n")


if __name__ == "__main__":
    asyncio.run(main())
